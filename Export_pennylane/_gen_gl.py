# -*- coding: utf-8 -*-
import openpyxl
from collections import defaultdict
SRC = r'C:\Users\guill\Code\TM-Budget\Export_pennylane\Provisoire_PENNYLANE_TERRA_MUCHO_Grand_livre_(2025_01_01_2025_12_31) (1).xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['Grand livre']

VALID = {'1.1.1','1.1.2','1.1.3','1.1.4','1.2.1','1.2.2','1.2.3','1.2.4','1.2.5','1.2.6','1.2.7','1.2.8',
'1.3.1','1.3.2','1.3.3','1.3.4','1.3.5','1.3.6','1.3.7','1.3.8','1.4.1','1.4.2','1.4.3','1.4.4','1.4.5','1.4.6','1.4.7',
'2.1.1','2.1.3','2.1.4','2.3.2','2.3.3','2.3.4','2.3.5','2.3.6','2.3.7','2.3.8','2.3.9','2.3.10','2.3.11','2.3.12','2.3.13',
'2.4.1','2.4.2','2.4.3','2.4.4','2.5.1','2.5.2','2.5.3','2.5.4','2.5.5','2.5.6','2.5.7','2.6.1','2.6.2','2.6.3',
'2.7.1','2.7.2','2.8.1','2.8.2','2.8.3','3.1.1','3.1.2','3.1.3','3.1.4','3.2.1','3.2.2','3.2.3','3.2.5',
'3.3.1','3.3.2','3.3.3','3.4.1','3.4.2','3.4.3','3.4.4','3.4.5','3.5.1','3.5.3','3.6.1','3.6.3','3.6.4','3.6.5','3.6.6','3.6.7','4.1.1'}

acct_default = {
 '604000':'2.4.1','606300':'1.4.3','606400':'1.4.3','613200':'1.4.1','613500':'1.3.5','613510':'1.3.5',
 '616300':'1.3.8','621000':'1.3.4','622600':'1.3.7','622610':'1.3.4','622700':'1.3.8','623000':'3.4.1',
 '623600':'3.4.5','623810':'2.3.3','624100':'3.1.3','625100':'3.1.1','625600':'2.5.4','625700':'1.4.3',
 '626200':'1.4.4','627000':'1.3.2','627510':'1.3.2','627520':'1.3.2','658000':'1.3.2','666000':'1.3.3'}

kw = [
 ('emily wilson','1.1.3'),
 ('marion cremona','2.3.2'),
 ('fidele yobo','2.3.13'),('fidele yobo','2.3.13'),('yobo gouem','2.3.13'),
 ('myra dunoyer','2.4.1'),('frais de mission de myra','2.4.1'),
 ('aline andree essono','2.1.4'),
 ('deborah bilau','2.1.3'),
 ('agence wooli','3.4.1'),
 ('marine ortuno','1.3.6'),
 ('shauri','1.3.6'),('moodle','2.4.4'),
 ('nikita loh','3.4.1'),('kayee au','3.4.1'),('mosaique','3.4.1'),
 ('kumquat','3.4.2'),('eric blanchet','2.5.3'),
 ('siteground','3.4.4'),('website','3.4.4'),
 ('rodec','1.3.4'),('interim assistance','1.3.4'),
 ('ascom','1.3.7'),('baker tilly','1.3.7'),
 ('retraite be','2.3.11'),('atelier retraite','2.3.11'),
 ('participant','2.3.7'),('cofe','2.3.7'),('societe immobiliere','2.3.3'),
 ('hotel ilomba','2.3.3'),('bracelet','2.3.11'),('journaux participantes','2.3.11'),
 ('team meeting','3.2.2'),('bxlle','3.2.2'),('bxelle','3.2.2'),('bruxelles','3.2.2'),
 ('nkolandom','2.5.4'),('visite','2.5.4'),
 ('ausha','2.8.2'),('descript','2.8.2'),
 ('openai','1.3.5'),('chatgpt','1.3.5'),('zoom','1.3.5'),('mailchimp','1.3.5'),('smartsuite','1.3.5'),
 ('odoo','1.3.5'),('pennylane','1.3.5'),('funds','1.3.5'),
 ('chargeur','1.2.1'),('laptop','1.2.1'),('couverts','1.2.6'),('assiettes','1.2.6'),
 ('henry monnier','1.4.1'),('yeleh','1.4.1'),('airbnb','1.4.1'),
 ('free mobile','1.4.4'),('internet','1.4.4'),('telephone','1.4.4'),
 ('visa','1.3.8'),('rapatriement','1.3.8'),('rappatriement','1.3.8'),
 ('comm. transfert','1.3.2'),('frais de correspondant','1.3.2'),('credit agricole','1.3.2'),
 ('comm de transfert','1.3.2'),('access','1.3.2'),('commission de change','1.3.2'),
 ('lettrage','1.3.3'),
]

def norm(s):
    s = (s or '').lower()
    for a,b in [('é','e'),('è','e'),('ê','e'),('à','a'),('ô','o'),('î','i'),('ï','i'),('û','u'),('ç','c')]:
        s = s.replace(a,b)
    return s

def pick(acct, lab):
    L = norm(lab)
    for k,c in kw:
        if k in L:
            return c
    return acct_default.get(acct,'4.1.1')

rows=[]; bylb=defaultdict(lambda:[0,0.0]); skipped7=0
for r in range(2, ws.max_row+1):
    no = ws.cell(r,1).value
    if no is None: continue
    no = str(no)[:6]
    if no.startswith('7'): skipped7+=1; continue
    d = ws.cell(r,6).value or 0; c = ws.cell(r,7).value or 0
    net = round(float(d)-float(c),2)
    if net == 0: continue
    dt = ws.cell(r,3).value
    raw = ws.cell(r,5).value or ''
    lab = raw.replace("'","''")
    ds = dt.strftime('%Y-%m-%d')
    code = pick(no, raw)
    assert code in VALID, code
    rows.append((ds,net,lab,code))
    bylb[code][0]+=1; bylb[code][1]+=net

L=[]
L.append("-- GL reel 2025 (depenses) importe depuis Pennylane Grand livre")
L.append("-- Recettes (comptes 7xxxxx) exclues. Montant = Debit - Credit (net).")
L.append("-- line_id resolu via structure_lines.code. Transactionnel.")
L.append("begin;")
L.append("delete from gl_entries where entry_type='Dépense' and entry_date >= '2025-01-01' and entry_date <= '2025-12-31';")
L.append("insert into gl_entries (entry_date, entry_type, label, amount, line_id)")
L.append("select v.d::date, 'Dépense', v.lab, v.amt, sl.id")
L.append("from (values")
vals=[f"  ('{ds}','{lab}',{net},'{code}')" for ds,net,lab,code in rows]
L.append(",\n".join(vals))
L.append(") as v(d,lab,amt,code)")
L.append("join structure_lines sl on sl.code = v.code;")
L.append("commit;")
sql="\n".join(L)
open(r'C:\Users\guill\Code\TM-Budget\supabase\seed_gl_2025.sql','w',encoding='utf-8').write(sql)

tot=sum(n for _,n,_,_ in rows)
print('rows inserted:',len(rows),'| skipped recettes(7x):',skipped7,'| TOTAL net dep=',round(tot,2))
print('--- per LB (desc) ---')
for c in sorted(bylb,key=lambda x:-bylb[x][1]):
    print(f'{c:7} n={bylb[c][0]:3} {round(bylb[c][1],2)}')
